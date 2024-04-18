import os
import glob
import xlrd
from openpyxl import Workbook

def convert_xls_to_xlsx(folder_path):
    xls_files = glob.glob(os.path.join(folder_path, '*.xls'))

    for xls_file in xls_files:
        # Open the xls file
        xls_workbook = xlrd.open_workbook(xls_file)
        
        # Create a new xlsx file
        xlsx_file = os.path.splitext(xls_file)[0] + '.xlsx'
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)  # Remove default sheet

        for sheet_name in xls_workbook.sheet_names():
            xls_sheet = xls_workbook.sheet_by_name(sheet_name)
            new_sheet = new_workbook.create_sheet(title=sheet_name)

            for row_index in range(xls_sheet.nrows):
                row = xls_sheet.row(row_index)
                for col_index, cell in enumerate(row):
                    new_sheet.cell(row=row_index+1, column=col_index+1).value = cell.value

        # Save xlsx file
        new_workbook.save(filename=xlsx_file)
        print(f"{xls_file} converted to {xlsx_file}")

# Example usage
folder_path = 'C:/Anlise_dados/anlise/XLS'
convert_xls_to_xlsx(folder_path)
